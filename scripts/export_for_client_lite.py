"""
Wersja dla klienta — tylko Artykuly + Kategorie.
Zachowuje oryginalny Arkusz1 nietkniety.
"""
import sys
import os
sys.path.append(os.path.join(os.path.dirname(__file__), '..', 'backend'))

import shutil
from pathlib import Path

import pymysql
from openpyxl import load_workbook

import export_to_unify as full  # reuzyj logiki

SRC = Path(r"c:\projects\repos\RaoApp\temp\Asortyment - Produkty - Maszyny - Toolsmart.xlsx")
DST = Path(r"c:\projects\repos\RaoApp\temp\Ujednolicenie_Artykuly_Kategorie.xlsx")

# wybieramy tylko 2 arkusze
WANTED = ["01_Artykuly", "02_Kategorie"]
SHEETS = [s for s in full.SHEETS if s["name"] in WANTED]


def write_simple_instruction(wb):
    ws = wb.create_sheet("00_Instrukcja", 0)
    from openpyxl.styles import Font, PatternFill, Alignment

    TITLE = Font(bold=True, color="FFFFFF", size=13)
    FILL = PatternFill("solid", fgColor="305496")
    H2 = Font(bold=True, size=12)
    RED = Font(bold=True, size=12, color="C00000")

    lines = [
        ("Ujednolicenie — Artykuly i Kategorie", TITLE, FILL),
        ("", None, None),
        ("CEL:", H2, None),
        ("   Uporzadkowanie slownikow artykulow/uslug oraz kategorii.", None, None),
        ("   Uzupelnienie brakujacych numerow wewnetrznych, marek, modeli, opisow.", None, None),
        ("   Po edycji — plik wraca do nas i zostanie zaimportowany z powrotem do bazy.", None, None),
        ("", None, None),
        ("JAK EDYTOWAC:", H2, None),
        ("   1. NIE zmieniaj kolumny 'ID (nie zmieniac)' — to klucz z bazy.", None, None),
        ("   2. W kolumnie 'Akcja' wybierz z listy:", None, None),
        ("        (puste)  — nie zmieniam tego wiersza", None, None),
        ("        UPDATE   — zmieniam dane (zostaw ID)", None, None),
        ("        DELETE   — usun ten artykul/kategorie z bazy", None, None),
        ("        NEW      — nowy wiersz: zostaw ID puste, wypelnij pola", None, None),
        ("   3. Zielone kolumny (Ile razy uzyty, Liczba artykulow, Data utworzenia)", None, None),
        ("      sa tylko do odczytu — informacyjnie.", None, None),
        ("   4. W kolumnie 'Uwagi klienta' mozesz zostawic dowolne notatki.", None, None),
        ("", None, None),
        ("ARKUSZE W PLIKU:", H2, None),
        ("   Arkusz1         — Twoj oryginal z notatkami (nieruszany)", None, None),
        ("   01_Artykuly     — 397 artykulow i uslug do ujednolicenia", None, None),
        ("   02_Kategorie    — 22 kategorie (ze statystyka liczby artykulow)", None, None),
        ("", None, None),
        ("GLOWNE BRAKI DO UZUPELNIENIA:", RED, None),
        ("   - 397/397 artykulow bez numeru wewnetrznego (pole kluczowe dla ewidencji!)", None, None),
        ("   - 397/397 bez opisu i numeru rejestracyjnego", None, None),
        ("   - 386/397 bez marki, 390/397 bez modelu i numeru seryjnego", None, None),
        ("   - 69 artykulow nigdy nie uzytych w umowie (kandydaci do usuniecia)", None, None),
        ("", None, None),
        ("DUPLIKATY (63 grupy, lacznie ~186 wierszy nadmiarowych):", H2, None),
        ("   Artykuly o tej samej nazwie to:", None, None),
        ("     a) albo OSOBNE FIZYCZNE SZTUKI tego samego typu (19 grup) —", None, None),
        ("        rozroznij je numerem wewnetrznym/seryjnym;", None, None),
        ("     b) albo PRAWDZIWE DUPLIKATY (44 grupy) do scalenia —", None, None),
        ("        zostaw jeden z UPDATE, reszte oznacz DELETE.", None, None),
        ("", None, None),
        ("KONWENCJE:", H2, None),
        ("   Pola 1/0     : wpisz 1 (tak) lub 0 (nie), nie 'tak'/'nie'", None, None),
        ("   Kwoty        : liczba z kropka dziesietna (np. 1234.56), bez 'PLN'", None, None),
        ("   Daty         : format YYYY-MM-DD (np. 2026-04-22)", None, None),
        ("   ID kategorii : liczba z arkusza 02_Kategorie", None, None),
    ]
    for i, (txt, font, fill) in enumerate(lines, 1):
        c = ws.cell(row=i, column=1, value=txt)
        if font:
            c.font = font
        if fill:
            c.fill = fill
            c.alignment = Alignment(vertical="center", indent=1)
            ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=6)
            ws.row_dimensions[i].height = 22
    ws.column_dimensions["A"].width = 100


def main():
    if DST.exists():
        DST.unlink()
    shutil.copy(SRC, DST)
    print(f"[1/3] Kopia: {DST.name}")

    wb = load_workbook(DST)
    existing = set(wb.sheetnames)

    conn = pymysql.connect(
        host=full.DB_HOST, port=full.DB_PORT, user=full.DB_USER,
        password=full.DB_PASS, database=full.DB_NAME, charset="utf8mb4",
    )
    try:
        print("[2/3] Eksport arkuszy:")
        for spec in SHEETS:
            rows = full.fetch_rows(conn, spec["sql"])
            full.write_sheet(wb, spec, rows)
            print(f"       OK  {spec['name']:<18} {len(rows):>5} wierszy")
    finally:
        conn.close()

    # Instrukcja — na pierwsza pozycje
    write_simple_instruction(wb)
    # ruszaj do przodu
    while wb.sheetnames[0] != "00_Instrukcja":
        wb.move_sheet("00_Instrukcja", offset=-1)

    # porzadek koncowy: 00_Instrukcja, Arkusz1, 01_Artykuly, 02_Kategorie
    desired = ["00_Instrukcja", "Arkusz1", "01_Artykuly", "02_Kategorie"]
    for i, name in enumerate(desired):
        if name in wb.sheetnames:
            cur = wb.sheetnames.index(name)
            if cur != i:
                wb.move_sheet(name, offset=(i - cur))

    wb.save(DST)
    print(f"[3/3] Zapisano: {DST}")
    print(f"       Arkusze: {wb.sheetnames}")


if __name__ == "__main__":
    main()
