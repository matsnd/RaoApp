"""
Export danych słownikowych z bazy rao_new do pliku Excel
dla klienta do ujednolicenia, uzupełnienia i późniejszego re-importu.

Zasady:
  - Oryginalny Arkusz1 klienta NIE jest ruszany — dopisujemy nowe arkusze do istniejącego pliku.
  - Każdy arkusz = jedna encja DB (articles, categories, contractors, ...)
  - Kolumna A = "ID (nie zmieniać)" — klucz główny z bazy. Puste ID = nowy rekord.
  - Kolumna B = "Akcja" — {'', 'UPDATE', 'DELETE', 'NEW'} (walidacja dropdown)
  - Nagłówki = czytelne polskie nazwy. Pełne mapowanie → spec/24_EXPORT_UJEDNOLICENIE.md
  - Statystyki użycia (liczba umów / liczba artykułów) pokazują co jest ważne.
  - Zamrożony pierwszy wiersz, autofilter, szerokość kolumn, kolorowe nagłówki.

Uruchomienie: python export_to_unify.py
"""
import sys
import os
sys.path.append(os.path.join(os.path.dirname(__file__), '..', '..', 'backend'))

from __future__ import annotations

import shutil
from pathlib import Path

import pymysql
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# -------------------- konfiguracja --------------------
DB_HOST = "localhost"
DB_PORT = 3306
DB_USER = "rao_user"
DB_PASS = "<<DB_PASSWORD_PLACEHOLDER>>"
DB_NAME = "rao_new"

SOURCE_XLSX = Path(r"c:\projects\repos\RaoApp\temp\Asortyment - Produkty - Maszyny - Toolsmart.xlsx")
OUT_XLSX = Path(r"c:\projects\repos\RaoApp\temp\Export_do_ujednolicenia.xlsx")

# -------------------- styl --------------------
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ID_FILL = PatternFill("solid", fgColor="F2F2F2")
ID_FONT = Font(bold=True, color="7F7F7F")
ACTION_FILL = PatternFill("solid", fgColor="FFF2CC")
STATS_FILL = PatternFill("solid", fgColor="E2EFDA")
TITLE_FILL = PatternFill("solid", fgColor="305496")
TITLE_FONT = Font(bold=True, color="FFFFFF", size=13)
LEGEND_FILL = PatternFill("solid", fgColor="FFF2CC")
LEGEND_FONT = Font(italic=True, color="595959", size=10)
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# -------------------- definicje arkuszy --------------------
# Każdy arkusz:
#   name, title (tytuł nad nagłówkiem), sql (zapytanie), columns: [(nagłówek, klucz_dict, szerokość)]
#   transforms: {klucz: lambda row→wartość}  — kolumny wyliczane
#   stats_cols: lista kluczy które pokazać kolorem "stat" (read-only)
#   legend: lista linii pomocniczych (pod nagłówkiem)

SHEETS: list[dict] = [
    {
        "name": "01_Artykuly",
        "title": "Artykuły i usługi — słownik przedmiotów wynajmu (397)",
        "sql": """
            SELECT a.id, a.name, a.is_service, a.internal_number, a.registration_no,
                   a.serial_no, a.brand, a.model, a.replacement_value,
                   c.name AS category_name, a.category_id,
                   ow.name AS owner_name, a.owner_id,
                   b.name AS branch_name, a.branch_id,
                   a.description, a.notes, a.rental_days, a.article_type,
                   (SELECT COUNT(*) FROM contract_positions cp WHERE cp.article_id=a.id) AS usage_count,
                   DATE_FORMAT(a.created_at, '%%Y-%%m-%%d') AS created_at
            FROM articles a
            LEFT JOIN categories c ON c.id=a.category_id
            LEFT JOIN contractors ow ON ow.id=a.owner_id
            LEFT JOIN branches b ON b.id=a.branch_id
            ORDER BY a.is_service, c.name, a.name
        """,
        "columns": [
            ("ID (nie zmieniac)", "id", 10),
            ("Akcja", "_action", 12),
            ("Nazwa", "name", 45),
            ("Rodzaj (artykul/usluga)", "_kind", 18),
            ("Numer wewnetrzny", "internal_number", 18),
            ("Numer rejestracyjny", "registration_no", 18),
            ("Numer seryjny", "serial_no", 18),
            ("Marka", "brand", 15),
            ("Model", "model", 15),
            ("Wartosc odtworzeniowa (PLN)", "replacement_value", 18),
            ("Kategoria", "category_name", 30),
            ("ID kategorii", "category_id", 10),
            ("Wlasciciel", "owner_name", 28),
            ("ID wlasciciela", "owner_id", 10),
            ("Oddzial", "branch_name", 18),
            ("ID oddzialu", "branch_id", 10),
            ("Opis", "description", 40),
            ("Notatki wewnetrzne", "notes", 30),
            ("Domyslna liczba dni wynajmu", "rental_days", 14),
            ("Typ artykulu (tag)", "article_type", 18),
            ("Ile razy uzyty w umowach", "usage_count", 14),
            ("Data utworzenia", "created_at", 12),
            ("Uwagi klienta", "_notes", 40),
        ],
        "transforms": {
            "_kind": lambda r: "usluga" if r["is_service"] else "artykul",
        },
        "stats_cols": ["usage_count", "created_at"],
        "legend": [
            "DIAGNOZA BRAKOW: 397/397 bez numeru wewnetrznego, bez opisu i bez numeru rejestracyjnego.",
            "386/397 bez marki, 390/397 bez modelu i numeru seryjnego. 69 artykulow nieuzywanych w zadnej umowie.",
            "ROZNIE: kolumna 'Rodzaj' wpisz 'artykul' lub 'usluga'. ID kategorii/wlasciciela/oddzialu — patrz arkusze 02/03/05.",
        ],
    },
    {
        "name": "02_Kategorie",
        "title": "Kategorie artykulow (22)",
        "sql": """
            SELECT c.id, c.name, c.code, c.description,
                   (SELECT COUNT(*) FROM articles a WHERE a.category_id=c.id) AS article_count
            FROM categories c ORDER BY article_count DESC, c.name
        """,
        "columns": [
            ("ID (nie zmieniac)", "id", 10),
            ("Akcja", "_action", 12),
            ("Nazwa kategorii", "name", 40),
            ("Kod (skrot)", "code", 14),
            ("Opis", "description", 50),
            ("Liczba artykulow", "article_count", 14),
            ("Uwagi klienta", "_notes", 40),
        ],
        "stats_cols": ["article_count"],
        "legend": [
            "Klient w Arkusz1 sygnalizuje reorganizacje kategorii. Obecny model jest PLASKI (brak parent_id).",
            "Jezeli potrzebna hierarchia (3-poziomowa: glowna → podkategoria → typ) — zglos do dewelopera (zmiana schematu DB).",
        ],
    },
    {
        "name": "03_Kontrahenci",
        "title": "Kontrahenci — klienci i dostawcy (581)",
        "sql": """
            SELECT c.id, c.name, c.name_short, c.nip, c.regon, c.pesel,
                   c.postal_code, c.city, c.street, c.unit,
                   c.email, c.contact_person1, c.phone1, c.contact_person2, c.phone2,
                   c.landline_phone, c.website, c.is_supplier, c.notes,
                   (SELECT COUNT(*) FROM contracts x WHERE x.contractor_id=c.id) AS contract_count,
                   DATE_FORMAT(c.created_at, '%%Y-%%m-%%d') AS created_at
            FROM contractors c ORDER BY contract_count DESC, c.name
        """,
        "columns": [
            ("ID (nie zmieniac)", "id", 10),
            ("Akcja", "_action", 12),
            ("Nazwa pelna", "name", 50),
            ("Nazwa skrocona", "name_short", 30),
            ("NIP", "nip", 14),
            ("REGON", "regon", 14),
            ("PESEL", "pesel", 14),
            ("Kod pocztowy", "postal_code", 10),
            ("Miejscowosc", "city", 20),
            ("Ulica", "street", 30),
            ("Lokal", "unit", 10),
            ("E-mail", "email", 30),
            ("Osoba kontaktowa 1", "contact_person1", 25),
            ("Telefon 1", "phone1", 18),
            ("Osoba kontaktowa 2", "contact_person2", 25),
            ("Telefon 2", "phone2", 18),
            ("Telefon stacjonarny", "landline_phone", 18),
            ("Strona WWW", "website", 25),
            ("Czy dostawca (1/0)", "is_supplier", 12),
            ("Notatki", "notes", 40),
            ("Liczba umow", "contract_count", 12),
            ("Data utworzenia", "created_at", 12),
            ("Uwagi klienta", "_notes", 40),
        ],
        "stats_cols": ["contract_count", "created_at"],
        "legend": [
            "DIAGNOZA BRAKOW: 567/581 bez e-maila, 348/581 bez telefonu, 51 bez ulicy, 31 bez miasta, 8 bez NIP.",
            "Czy dostawca: 1 = dostawca sprzetu, 0 = najemca. Obecnie WSZYSCY (581) sa najemcami — jesli masz dostawcow, oznacz 1.",
            "Liczba umow: 0 umow = kandydat do DELETE.",
        ],
    },
    {
        "name": "04_Adresy_dostawy",
        "title": "Adresy dostawy — przypisane do kontrahentow (280)",
        "sql": """
            SELECT ca.id, ca.contractor_id, c.name AS contractor_name,
                   ca.name, ca.country_code, ca.postal_code, ca.city, ca.street,
                   ca.contact_person, ca.phone, ca.email,
                   ca.is_default_delivery, ca.is_headquarters,
                   ca.latitude, ca.longitude, ca.notes,
                   DATE_FORMAT(ca.created_at, '%%Y-%%m-%%d') AS created_at
            FROM contractor_addresses ca
            LEFT JOIN contractors c ON c.id=ca.contractor_id
            ORDER BY c.name, ca.is_headquarters DESC, ca.is_default_delivery DESC
        """,
        "columns": [
            ("ID (nie zmieniac)", "id", 10),
            ("Akcja", "_action", 12),
            ("ID kontrahenta", "contractor_id", 12),
            ("Kontrahent (nazwa)", "contractor_name", 40),
            ("Opis adresu", "name", 30),
            ("Kraj", "country_code", 8),
            ("Kod pocztowy", "postal_code", 10),
            ("Miejscowosc", "city", 20),
            ("Ulica", "street", 30),
            ("Osoba kontaktowa", "contact_person", 25),
            ("Telefon", "phone", 18),
            ("E-mail", "email", 25),
            ("Adres domyslny dostawy (1/0)", "is_default_delivery", 14),
            ("Siedziba firmy (1/0)", "is_headquarters", 12),
            ("Szerokosc geograf.", "latitude", 14),
            ("Dlugosc geograf.", "longitude", 14),
            ("Notatki", "notes", 30),
            ("Data utworzenia", "created_at", 12),
            ("Uwagi klienta", "_notes", 40),
        ],
        "stats_cols": ["created_at"],
        "legend": [
            "Kazdy kontrahent moze miec wiele adresow (magazyny, place budowy, siedziba).",
            "Wspolrzedne geograficzne: uzywane przez mape Nominatim. Zostaw puste jezeli nie znasz.",
        ],
    },
    {
        "name": "05_Oddzialy",
        "title": "Oddzialy wlasnej firmy (2)",
        "sql": "SELECT id, name, address, postal_code, city, street, created_at FROM branches ORDER BY id",
        "columns": [
            ("ID (nie zmieniac)", "id", 10),
            ("Akcja", "_action", 12),
            ("Nazwa oddzialu", "name", 30),
            ("Adres (jedno pole)", "address", 40),
            ("Kod pocztowy", "postal_code", 10),
            ("Miejscowosc", "city", 20),
            ("Ulica", "street", 30),
            ("Data utworzenia", "created_at", 12),
            ("Uwagi klienta", "_notes", 40),
        ],
        "stats_cols": ["created_at"],
        "legend": [
            "Oddzialy sluza do podzialu statystyk i sprzedazy w dashboardzie.",
        ],
    },
    {
        "name": "06_Handlowcy",
        "title": "Handlowcy — wystawcy umow (3)",
        "sql": "SELECT id, name, phone, is_active, commission_rate FROM salespeople ORDER BY name",
        "columns": [
            ("ID (nie zmieniac)", "id", 10),
            ("Akcja", "_action", 12),
            ("Imie i nazwisko", "name", 30),
            ("Telefon", "phone", 20),
            ("Aktywny (1/0)", "is_active", 10),
            ("Prowizja (%)", "commission_rate", 12),
            ("Uwagi klienta", "_notes", 40),
        ],
        "legend": [
            "Aktywny=0: handlowiec jest ukryty w listach, ale historia umow zachowana.",
            "Prowizja: procent od wartosci umowy. Uzywany w statystykach w dashboardzie.",
        ],
    },
    {
        "name": "07_Typy_stawek",
        "title": "Typy stawek — rozliczanie umow (3)",
        "sql": "SELECT id, name, description, is_dependent FROM rate_types ORDER BY id",
        "columns": [
            ("ID (nie zmieniac)", "id", 10),
            ("Akcja", "_action", 12),
            ("Nazwa", "name", 40),
            ("Opis", "description", 60),
            ("Zalezny od pogody/innych (1/0)", "is_dependent", 14),
            ("Uwagi klienta", "_notes", 40),
        ],
        "legend": [
            "Typy stawek uzywane w warunkach rozliczania pozycji umowy (np. 'dobowa', 'godzinowa').",
        ],
    },
    {
        "name": "08_Presety_oplat",
        "title": "Grupy presetow oplat serwisowych (2)",
        "sql": """
            SELECT fpg.id, fpg.name, fpg.contract_type, fpg.description,
                   fpg.is_default, fpg.sort_order,
                   (SELECT COUNT(*) FROM service_fee_templates t WHERE t.preset_id=fpg.id) AS templates_count
            FROM fee_preset_groups fpg ORDER BY fpg.contract_type, fpg.sort_order
        """,
        "columns": [
            ("ID (nie zmieniac)", "id", 10),
            ("Akcja", "_action", 12),
            ("Nazwa presetu", "name", 35),
            ("Typ umowy (S=najem / U=usluga)", "contract_type", 16),
            ("Opis", "description", 50),
            ("Domyslny (1/0)", "is_default", 10),
            ("Kolejnosc", "sort_order", 10),
            ("Liczba szablonow", "templates_count", 14),
            ("Uwagi klienta", "_notes", 40),
        ],
        "stats_cols": ["templates_count"],
        "legend": [
            "Preset grupuje szablony oplat. Kazda nowa umowa dostaje kopie szablonow z wybranego presetu.",
        ],
    },
    {
        "name": "09_Szablony_oplat",
        "title": "Szablony oplat serwisowych (10)",
        "sql": """
            SELECT sft.id, fpg.id AS preset_id, fpg.name AS preset_name,
                   sft.contract_type, sft.sort_order, sft.name,
                   sft.amount_from, sft.amount_to, sft.unit, sft.description, sft.is_active
            FROM service_fee_templates sft
            LEFT JOIN fee_preset_groups fpg ON fpg.id=sft.preset_id
            ORDER BY sft.contract_type, sft.sort_order
        """,
        "columns": [
            ("ID (nie zmieniac)", "id", 10),
            ("Akcja", "_action", 12),
            ("ID presetu", "preset_id", 10),
            ("Preset (nazwa)", "preset_name", 30),
            ("Typ umowy (S/U)", "contract_type", 12),
            ("Kolejnosc", "sort_order", 10),
            ("Nazwa oplaty", "name", 35),
            ("Kwota od (PLN)", "amount_from", 14),
            ("Kwota do (PLN)", "amount_to", 14),
            ("Jednostka", "unit", 15),
            ("Opis", "description", 50),
            ("Aktywny (1/0)", "is_active", 10),
            ("Uwagi klienta", "_notes", 40),
        ],
        "legend": [
            "Kazdy szablon = jedna pozycja oplat. Uzywane jako wzorce przy zakladaniu nowej umowy.",
        ],
    },
    {
        "name": "10_Firma",
        "title": "Dane wlasnej firmy (1)",
        "sql": """
            SELECT id, name, name_short, nip, regon, postal_code, city, street,
                   bank_name, bank_account, header_text, numbering_start, increment_step
            FROM company
        """,
        "columns": [
            ("ID (nie zmieniac)", "id", 10),
            ("Akcja", "_action", 12),
            ("Nazwa pelna", "name", 40),
            ("Nazwa skrocona", "name_short", 25),
            ("NIP", "nip", 14),
            ("REGON", "regon", 14),
            ("Kod pocztowy", "postal_code", 10),
            ("Miejscowosc", "city", 20),
            ("Ulica", "street", 30),
            ("Bank", "bank_name", 30),
            ("Numer konta", "bank_account", 30),
            ("Naglowek na dokumentach", "header_text", 50),
            ("Numer poczatkowy umow", "numbering_start", 14),
            ("Krok numeracji", "increment_step", 12),
            ("Uwagi klienta", "_notes", 40),
        ],
        "legend": [
            "Dane wlasnej firmy pokazywane w naglowkach umow i protokolow.",
        ],
    },
]


# -------------------- pomocnicze --------------------
def fetch_rows(conn, sql: str) -> list[dict]:
    with conn.cursor(pymysql.cursors.DictCursor) as cur:
        cur.execute(sql)
        return list(cur.fetchall())


def apply_transforms(rows: list[dict], transforms: dict) -> list[dict]:
    if not transforms:
        return rows
    for r in rows:
        for key, fn in transforms.items():
            r[key] = fn(r)
    return rows


def _fmt(v):
    if v is None:
        return ""
    # Decimal/int pozostaw jako liczby — Excel sam sformatuje
    return v


def write_sheet(wb: Workbook, spec: dict, rows: list[dict]) -> None:
    ws = wb.create_sheet(spec["name"])

    # Rząd 1: tytuł arkusza (scalony)
    ws.cell(row=1, column=1, value=spec["title"]).font = TITLE_FONT
    ws.cell(row=1, column=1).fill = TITLE_FILL
    ws.cell(row=1, column=1).alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(spec["columns"]))
    ws.row_dimensions[1].height = 22

    # Rzędy 2..N: legenda
    legend = spec.get("legend", [])
    next_row = 2
    for line in legend:
        c = ws.cell(row=next_row, column=1, value="  " + line)
        c.font = LEGEND_FONT
        c.fill = LEGEND_FILL
        c.alignment = Alignment(vertical="center", horizontal="left")
        ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=len(spec["columns"]))
        ws.row_dimensions[next_row].height = 16
        next_row += 1

    # Pusty wiersz separator
    next_row += 1
    header_row = next_row

    # Nagłówki
    stats_cols = set(spec.get("stats_cols", []))
    for col_idx, (human, key, width) in enumerate(spec["columns"], 1):
        c = ws.cell(row=header_row, column=col_idx, value=human)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        c.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[header_row].height = 32

    # Dane
    transforms = spec.get("transforms", {})
    rows = apply_transforms(rows, transforms)

    for r_off, row in enumerate(rows, 1):
        excel_row = header_row + r_off
        for col_idx, (human, key, width) in enumerate(spec["columns"], 1):
            if key is None or key == "_action" or key == "_notes":
                value = ""  # do uzupełnienia przez klienta
            else:
                value = row.get(key, "")
            cell = ws.cell(row=excel_row, column=col_idx, value=_fmt(value))
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=(width > 30))
            # kolorowanie
            if key == "id":
                cell.fill = ID_FILL
                cell.font = ID_FONT
            elif key == "_action":
                cell.fill = ACTION_FILL
            elif key in stats_cols:
                cell.fill = STATS_FILL

    # Walidacja "Akcja" — dropdown
    action_col_idx = next((i for i, (_, k, _) in enumerate(spec["columns"], 1) if k == "_action"), None)
    if action_col_idx and rows:
        dv = DataValidation(
            type="list",
            formula1='"UPDATE,DELETE,NEW"',
            allow_blank=True,
            showDropDown=False,  # pokazuj strzałkę
        )
        dv.error = "Dopuszczalne: UPDATE, DELETE, NEW lub puste."
        dv.errorTitle = "Nieprawidlowa akcja"
        letter = get_column_letter(action_col_idx)
        dv.add(f"{letter}{header_row + 1}:{letter}{header_row + len(rows) + 500}")  # zapas na nowe wiersze
        ws.add_data_validation(dv)

    # Freeze panes: pod nagłówkiem + kolumna B (ID + Akcja)
    ws.freeze_panes = ws.cell(row=header_row + 1, column=3)

    # Autofilter tylko na nagłówku + dane
    if rows:
        last_col_letter = get_column_letter(len(spec["columns"]))
        ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{header_row + len(rows)}"


def write_instruction_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("00_Instrukcja", 0)

    lines = [
        ("Export danych do ujednolicenia", TITLE_FONT, TITLE_FILL),
        ("", None, None),
        ("CEL:", Font(bold=True, size=12), None),
        ("   Uporzadkowanie slownikow (artykuly, kontrahenci, kategorie, oddzialy, ...)", None, None),
        ("   Uzupelnienie brakujacych pol (numery wewnetrzne, email, telefon, kategorie).", None, None),
        ("   Zmiany zostana pozniej zaimportowane z powrotem do bazy.", None, None),
        ("", None, None),
        ("JAK EDYTOWAC:", Font(bold=True, size=12), None),
        ("   1. NIE zmieniaj wartosci w kolumnie 'ID (nie zmieniac)' — to klucz z bazy.", None, None),
        ("   2. W kolumnie 'Akcja' wpisz jedna z wartosci (dropdown):", None, None),
        ("        (puste) — brak zmian, wiersz zostanie zignorowany przy imporcie", None, None),
        ("        UPDATE  — zaktualizuj wiersz wg wpisanych danych", None, None),
        ("        DELETE  — usun wiersz z bazy (zastanow sie — historyczne powiazania!)", None, None),
        ("        NEW     — nowy rekord: zostaw pusty 'ID' i wypelnij pozostale kolumny", None, None),
        ("   3. Kolumny zielone (statystyki typu 'Ile razy uzyty', 'Liczba umow') sa TYLKO DO ODCZYTU.", None, None),
        ("   4. Kolumna 'Uwagi klienta' na koncu arkusza — wpisuj komentarze dla dewelopera.", None, None),
        ("   5. Kazdy arkusz ma wlasna legende (zolte wiersze pod tytulem).", None, None),
        ("", None, None),
        ("ARKUSZE W PLIKU:", Font(bold=True, size=12), None),
        ("   Arkusz1                 — Twoj oryginal (nieruszany, notatki do artykulow)", None, None),
        ("   01_Artykuly             — 397 artykulow i uslug", None, None),
        ("   02_Kategorie            — 22 kategorie artykulow", None, None),
        ("   03_Kontrahenci          — 581 klientow i potencjalnych dostawcow", None, None),
        ("   04_Adresy_dostawy       — 280 adresow (magazyny, place budowy)", None, None),
        ("   05_Oddzialy             — 2 oddzialy wlasnej firmy", None, None),
        ("   06_Handlowcy            — 3 handlowcow wystawiajacych umowy", None, None),
        ("   07_Typy_stawek          — 3 typy rozliczania (dobowa, godzinowa, ...)", None, None),
        ("   08_Presety_oplat        — 2 presety oplat serwisowych", None, None),
        ("   09_Szablony_oplat       — 10 szablonow oplat (wzorce do nowych umow)", None, None),
        ("   10_Firma                — dane wlasnej firmy (naglowek dokumentow)", None, None),
        ("", None, None),
        ("GLOWNE BRAKI W DANYCH (do uzupelnienia):", Font(bold=True, size=12, color="C00000"), None),
        ("   Artykuly:  397/397 bez numeru wewnetrznego, bez opisu i bez numeru rejestracyjnego", None, None),
        ("              386/397 bez marki, 390/397 bez modelu i numeru seryjnego", None, None),
        ("              69 artykulow NIGDY nie uzytych w umowie (kandydaci do DELETE)", None, None),
        ("   Kontrahenci: 567/581 bez e-maila, 348/581 bez telefonu, 51 bez ulicy", None, None),
        ("              8 bez NIP, 31 bez miasta", None, None),
        ("   Kategorie: plaska struktura — klient sygnalizuje potrzebe hierarchii", None, None),
        ("", None, None),
        ("KONWENCJE:", Font(bold=True, size=12), None),
        ("   - Pola logiczne wpisuj jako 1 (tak) lub 0 (nie).", None, None),
        ("   - Kwoty — liczby, bez spacji i waluty (np. 1234.56).", None, None),
        ("   - Daty — format YYYY-MM-DD (np. 2026-04-22).", None, None),
        ("   - ID kategorii / wlasciciela / oddzialu — wpisuj liczbe z odpowiedniego arkusza.", None, None),
        ("", None, None),
        ("DOKUMENTACJA TECHNICZNA:", Font(bold=True, size=12), None),
        ("   spec/24_EXPORT_UJEDNOLICENIE.md — pelne mapowanie kolumn na pola bazy (dla dewelopera)", None, None),
    ]

    for i, (text, font, fill) in enumerate(lines, 1):
        c = ws.cell(row=i, column=1, value=text)
        if font:
            c.font = font
        if fill:
            c.fill = fill
            c.alignment = Alignment(vertical="center", indent=1)
            ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=6)
            ws.row_dimensions[i].height = 22

    ws.column_dimensions["A"].width = 100


def main():
    # 1. Skopiuj źródłowy plik do output (nie nadpisujemy oryginału klienta)
    if OUT_XLSX.exists():
        OUT_XLSX.unlink()
    shutil.copy(SOURCE_XLSX, OUT_XLSX)
    print(f"[1/3] Kopia pliku: {OUT_XLSX.name}")

    # 2. Otwórz docelowy plik i dołóż arkusze
    wb = load_workbook(OUT_XLSX)
    existing = set(wb.sheetnames)
    print(f"       Istniejace arkusze (nie ruszane): {existing}")

    conn = pymysql.connect(
        host=DB_HOST, port=DB_PORT, user=DB_USER, password=DB_PASS,
        database=DB_NAME, charset="utf8mb4",
    )
    try:
        print("[2/3] Eksport arkuszy...")
        for spec in SHEETS:
            if spec["name"] in existing:
                print(f"       SKIP {spec['name']} (juz istnieje)")
                continue
            rows = fetch_rows(conn, spec["sql"])
            write_sheet(wb, spec, rows)
            print(f"       OK   {spec['name']:<22} {len(rows):>5} wierszy")

        # instrukcja jako pierwszy arkusz
        if "00_Instrukcja" not in existing:
            write_instruction_sheet(wb)
            # przesun instrukcje na pierwsza pozycje
            idx = wb.sheetnames.index("00_Instrukcja")
            wb.move_sheet("00_Instrukcja", offset=-idx)
            print("       OK   00_Instrukcja           — instrukcja dla klienta")
    finally:
        conn.close()

    # 3. Zapisz
    wb.save(OUT_XLSX)
    print(f"[3/3] Zapisano: {OUT_XLSX}")
    print()
    print("GOTOWE. Przeslij klientowi plik:")
    print(f"   {OUT_XLSX}")


if __name__ == "__main__":
    main()
